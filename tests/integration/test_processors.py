"""Integration tests for document processors.

Tests the interaction between different processor types and validates
the complete processing pipeline functionality.
"""

import logging
from pathlib import Path
from typing import Generator

import pandas as pd
import pytest
from openpyxl import Workbook

from src.ml.pipeline import ExcelProcessor, PipelineConfig, TextProcessor, WordProcessor
from src.ml.pipeline.config.settings import ProcessingConfig


@pytest.fixture
def sample_files(tmp_path: Path) -> Generator[dict[str, Path], None, None]:
    """Create sample files of different types for testing.

    Args:
        tmp_path: Pytest fixture providing temporary directory

    Yields:
        Dict containing paths to sample files of each type
    """
    # Create text file
    text_file = tmp_path / "sample.txt"
    text_file.write_text("Sample text content\nWith multiple lines\n\nAnd paragraphs")

    # Create Excel file
    excel_file = tmp_path / "sample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Header"
    ws["A2"] = "Content"
    wb.save(excel_file)

    # Create Word file (mock)
    word_file = tmp_path / "sample.docx"
    word_file.write_bytes(b"Mock Word file content")

    yield {"text": text_file, "excel": excel_file, "word": word_file}


@pytest.fixture
def pipeline_config() -> PipelineConfig:
    """Create a pipeline configuration for testing.

    Returns:
        Configured PipelineConfig instance
    """
    return PipelineConfig(batch_size=10, max_retries=2, cache_enabled=False)


def test_processor_initialization(pipeline_config: PipelineConfig):
    """Test initialization of multiple processors with shared config."""
    processors = [
        TextProcessor(config=pipeline_config),
        ExcelProcessor(config=pipeline_config),
        WordProcessor(config=pipeline_config),
    ]

    for processor in processors:
        assert processor.config == pipeline_config
        assert not processor.is_initialized

        processor.initialize()
        assert processor.is_initialized

        processor.cleanup()
        assert not processor.is_initialized


def test_sequential_processing(sample_files: dict[str, Path], pipeline_config: PipelineConfig):
    """Test processing multiple file types in sequence."""
    processors = {
        "text": TextProcessor(config=pipeline_config),
        "excel": ExcelProcessor(config=pipeline_config),
        "word": WordProcessor(config=pipeline_config),
    }

    results = {}

    for file_type, processor in processors.items():
        with processor:
            results[file_type] = processor.process(sample_files[file_type])

        assert results[file_type]["metadata"]["file_type"] == f".{file_type}"
        assert "content" in results[file_type]
        assert processor.processed_files == {sample_files[file_type]}


def test_error_propagation(sample_files: dict[str, Path], caplog):
    """Test error handling and logging across processors."""
    caplog.set_level(logging.ERROR)

    # Test with invalid configuration
    with pytest.raises(ValueError, match="Invalid configuration"):
        TextProcessor(processing_config=ProcessingConfig(max_retries=-1))

    # Test with invalid file paths
    processor = TextProcessor()
    with processor:
        with pytest.raises(ValueError, match="File not found"):
            processor.process(Path("nonexistent.txt"))

    assert "Error processing file" in caplog.text


def test_resource_cleanup(sample_files: dict[str, Path], pipeline_config: PipelineConfig):
    """Test proper resource cleanup across multiple processors."""
    processors = [
        TextProcessor(config=pipeline_config),
        ExcelProcessor(config=pipeline_config),
        WordProcessor(config=pipeline_config),
    ]

    # Test cleanup after normal operation
    for processor in processors:
        with processor:
            assert processor.is_initialized
        assert not processor.is_initialized
        assert len(processor.processed_files) == 0

    # Test cleanup after exception
    processor = TextProcessor(config=pipeline_config)
    try:
        with processor:
            raise RuntimeError("Test exception")
    except RuntimeError:
        pass
    finally:
        assert not processor.is_initialized
        assert len(processor.processed_files) == 0
